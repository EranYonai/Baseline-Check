from difflib import SequenceMatcher

import openpyxl
from PyQt5 import QtWidgets, uic

import config
import sys
import traceback  # for debug purpose

# Global variables
name_to_color = {'red': '#FF0000', 'yellow': '#FFA500', 'green': '#008000', 'teal': '#74C2E1'}
Alph_Zero_Val = ord('A') - 1


class MainWindow(QtWidgets.QMainWindow):
    """
    Main dialog of the application
    """

    def __init__(self):
        """
        init - initialize on dialog call.
        """
        super(MainWindow, self).__init__()

        uic.loadUi(config.FILE_PATHS['MAIN_UI'], self)  # Loads MainWindow UI using pyqt5 uic.
        self.check_button.clicked.connect(self.scan_baseline)  # Check button click event
        self.actionRefresh_Database.triggered.connect(self.update_db)
        # Global self variables:
        self.wb = None
        self.excel_map = None
        # On initialization function calls:
        self.update_db()

    def scan_baseline(self):
        """
        Gets a string of a baseline description from GUI and colors the lines according to item's presence DB
        """
        try:
            row = None
            res = ''
            bl_text = self.textEdit.toPlainText()
            sheet_name = column_name = item_val = None
            extenders_flag = dongle_flag = rfg_flag = None

            for line in bl_text.split('\n'):
                if not line or line.startswith('-----'):  # New Section - reset variables
                    sheet_name = column_name = item_val = None
                elif sheet_name is not None:  # Item Category (sheet) was already found
                    #  Manipulation on the line before reading and splitting goes here
                    if ':' in line:
                        column_name, item_val = line.split(':')
                        # TODO: if line does not contain exactly 1 ':', error.
                    elif sheet_name == 'SPU':  # Temporary fix to SPU fields without ':'
                        line = line.replace('\t', ':')
                        column_name, item_val = line.split(':')
                    else: # in case there is no ':'
                        item_val = None
                    if item_val is not None:
                        item_val = item_val.strip()
                        column_name = column_name.strip()

                    #  ---- System special cases ----
                    if sheet_name == 'System':  # Columns in System WS
                        if column_name.startswith('Aquarium'):
                            column_name = 'Aquarium'
                    #  ---- WS special cases ----
                    elif sheet_name == 'WS':  # Columns in Workstation WS
                        pass
                        # TODO: ---->> if WS validation does funny things, the 'WS Service Tag' should be first, I need to fix this in AutoBaseline export. <<----
                    #  ---- Catheters special cases ----
                    elif sheet_name == 'Catheters':  # Columns in Catheters WS
                        if line.startswith('Extenders'):
                            extenders_flag = True
                        if line.startswith('Catheters'):
                            extenders_flag = None
                            # ^ This is to cover the case Catheters and Extenders are not divided by --------. ^
                        if extenders_flag is None and dongle_flag is None:
                            column_name = 'Catheters Catalog Number'
                            row = None
                        elif dongle_flag is None:
                            column_name = 'Extenders Catalog Number'
                            row = None

                        if line.startswith('Software Version'):
                            column_name = 'SW Version'
                        if line.startswith('Hardware Version'):
                            column_name = 'HW Version'
                    #  ---- Pacers special cases ----
                    elif sheet_name == 'Pacers':
                        if column_name.endswith('Serial Number'):
                            column_name = 'Pacer Serial Number'
                        # TODO: ---->> if pacer validation does funny things, the 'Serial Number' should be first, I need to fix this in AutoBaseline export. <<----
                    #  ---- UltraSound special cases ----
                    elif sheet_name == "UltraSound":
                        if column_name.endswith('Ultrasound System'):
                            column_name = 'Model'
                    # ---- RFGs special cases ----
                    elif sheet_name == 'RFGs':
                        if rfg_flag == 'stockert':
                            if column_name.startswith('Serial Number'):
                                column_name = 'Stockert SN'
                                # TODO: Change SN to be first in Auto Baseline export
                        if rfg_flag == 'smartablate':
                            if column_name.startswith('Serial Number'):
                                column_name = 'SMARTABLATE SN'
                            if column_name.startswith('System Software'):
                                column_name = "SW Version"
                        if rfg_flag == 'nmarq':
                            if column_name.startswith('SW Version'):
                                column_name = 'nMARQ SW'
                            if column_name.startswith('Serial Number'):
                                column_name = 'nMARQ SN'
                            if column_name.startswith('COOLFLOW pump S.N'):
                                column_name = 'Pump SN'
                            if column_name.startswith('COOLFLOW pump Model'):
                                column_name = 'Pump Model'
                                # TODO: Change SN to be first in Auto Baseline export

                #    ----------------------------------------------------------------------------------    #
                # Find which sheet to search in
                elif line.startswith('System #'):
                    sheet_name = 'System'
                    column_name = item_val = row = None
                elif line.startswith('Ultrasound #'):
                    sheet_name = 'UltraSound'
                    column_name = item_val = row = None
                elif line.startswith('Workstation #'):
                    sheet_name = 'WS'
                    column_name = item_val = row = None
                elif line.startswith('SPU'):
                    sheet_name = 'SPU'
                    # TODO: add ':' in Auto Baseline SPU export and remove first '|'.
                    column_name = item_val = row = None
                elif line.startswith(
                        'Extenders'):  # does not get here unless extenders and catheters are divided by '--'
                    sheet_name = 'Catheters'
                    column_name = 'Extenders Catalog Number'
                    extenders_flag = True
                    item_val = row = None
                elif line.startswith('Catheters'):
                    sheet_name = 'Catheters'
                    column_name = 'Catheters Catalog Number'
                    item_val = row = extenders_flag = dongle_flag = None
                elif line.startswith('Pacer') or line.startswith('Printer') or line.startswith('EPU'):
                    sheet_name = 'Pacers'
                    column_name = item_val = row = None
                elif line.startswith('qDOT Dongle'):
                    dongle_flag = True
                    sheet_name = 'Catheters'
                    column_name = item_val = row = None
                elif line.startswith('Stockert GmbH System RF Generator'):
                    sheet_name = 'RFGs'
                    rfg_flag = 'stockert'
                    print(rfg_flag)
                    column_name = item_val = row = None
                elif line.startswith('SMARTABLATE RF Generator'):
                    sheet_name = 'RFGs'
                    rfg_flag = 'smartablate'
                    column_name = item_val = row = None
                elif line.startswith('nGEN RF Generator'):
                    sheet_name = 'nGEN'
                    column_name = item_val = row = None
                elif line.startswith('nMARQ Multi Channel RF Generator'):
                    sheet_name = 'RFGs'
                    rfg_flag = 'nmarq'
                    column_name = item_val = row = None

                print(f'sheetname *{sheet_name}*, columnname *{column_name}*, itemval *{item_val}*')
                if item_val == "":  # in the case of "Extender:"
                    item_val = None

                # if Sheet, column and item_value have all been detected- Search for similarities in the specified column and sheet
                if sheet_name and column_name and sheet_name in self.excel_map and column_name in self.excel_map[
                    sheet_name] and item_val is not None:
                    if row is None:
                        answer = find_word_in_db(sheet_name=sheet_name, column=self.excel_map[sheet_name][column_name],
                                                 word=item_val)
                    else:
                        answer = find_word_in_db(sheet_name=sheet_name, column=self.excel_map[sheet_name][column_name],
                                                 word=item_val, row=row)
                    if answer is None:  # if cell wasn't found
                        res += f'{color_str("red", line)} --> Not found in EXCEL<br>'
                    else:
                        color = 'green' if answer['diff'] == 1 else 'yellow' if answer['diff'] >= 0.8 else 'red'
                        if color == 'green':  # Append colored line
                            res += f'{color_str(color, line)}<br>'  # Append colored line - green
                        else:  # Append colored line - if not green, show correction in teal
                            res += f'{color_str(color, line)} -->> {color_str("teal", answer["match"])}<br>'
                        print(f'found match in row: {answer["row"]}')
                        row = answer["row"]
                else:
                    res += f'{line}<br>'  # append unchecked lines as they are

            self.textEdit.setHtml(res)
        except Exception as e:
            print("Problem in scanning baselines: " + str(e))
            traceback.print_exc()

    def update_db(self):
        """
        Dumb functions for now, it is planned to open a child dialog and choose the db location.
        """
        print("should update db")
        self.wb = openpyxl.load_workbook(config.FILE_PATHS['EQP_EXCEL'])  # load excel workbook
        self.excel_map = explore_excel()


def color_str(color: str, word: str) -> str:
    """
    Adds color tags to the given color.

    :param color: str = 'green', 'red, 'yellow'.
    :param word: str of the word you want tagged.
    :return: string: 'color_tag + word + close_tag'
    :rtype: str
    """
    return f'<span style="color:{name_to_color[color]};">{word}</span>'


def find_word_in_db(sheet_name, column, word, row=None):
    """
    searches word in sheet name's columns

    :param sheet_name: string of the sheet's name
    :param column: the specific column in which to search
    :param word: the string to search
    :param row: the specific row, if known
    :return:
    :rtype: dict {'diff', 'row', 'match'}
    """

    def find_cell_in_db(column, row, word):
        """
        finds the word in given excel db, returns dict with information

        :param column: the specific column
        :param word: the string to search
        :param row: the specific row, if known
        :return:
        :rtype: dict {'diff', 'row', 'match'}
        """
        diff = 0
        cell_in_excel = str(ws[f'{column}{row}'].value)
        if cell_in_excel is not None:  # Iterate all rows of specified column
            diff = max([diff, similar(word, cell_in_excel.strip())])
            print(f'looking for *{word}* in row {str(row)}, current cell: *{cell_in_excel}*, diff is {str(diff)}')
            if diff == 1:  # Stops if perfect match was found
                return {'diff': diff, 'row': row, 'match': cell_in_excel}
        # if an exact match wasn't found, return the last cell in column? it actually returns None sometimes
        return {'diff': diff, 'row': row, 'match': cell_in_excel}

    # ----#
    try:
        ws = MainWindow.wb[sheet_name]
        if row is None:  # e.g. row unknown, searching row
            for row in range(3,
                             ws.max_row + 1):  # iterates all rows (row = tuple), added +1 because if the value was last it didn't get to it.
                ans = find_cell_in_db(column=column, row=row, word=word)
                if ans['diff'] == 1:
                    print(f"----found exact match for item in row {row}!!----")
                    return ans
        else:  # e.g. row is known
            return find_cell_in_db(column=column, row=row, word=word)

    except Exception as e:
        print("Problem finding word in db: " + str(e))


def similar(str1, str2):
    """
    Gets two string and returns their correlation in percentage (float 0-1).
    Using SequenceMatcher of difflib library.

    :param str1: String 1
    :param str2: String 2
    :return: float value
    """
    return SequenceMatcher(None, str1, str2).ratio()


def explore_excel():
    """
    Maps the excel file specified in config.py
    :return: dictionary mapping the excel {<sheet_title>: {<column_title>: <column_letter>, ...}, ...}
    """
    try:
        res = dict()
        MainWindow.wb = openpyxl.load_workbook(config.FILE_PATHS['EQP_EXCEL'])  # load excel workbook
        for sheet in MainWindow.wb:
            res[sheet.title] = dict()
            for cell in sheet[2]:  # Iterate cells in 2nd row
                res[sheet.title][cell.value] = chr(cell.column + Alph_Zero_Val)  # get column letter
        return res
    except Exception as e:
        print("Problem exploring excel: " + str(e))


if __name__ == "__main__":
    """It's boilerplate code that protects users from accidentally invoking the script when they didn't intend to. 
    Here are some common problems when the guard is omitted from a script: If you import the guardless script in 
    another script (e.g. import my_script_without_a_name_eq_main_guard), then the second script will trigger the 
    first to run at import time and using the second script's command line arguments. This is almost always a 
    mistake. If you have a custom class in the guardless script and save it to a pickle file, then unpickling it in 
    another script will trigger an import of the guardless script, with the same problems outlined in the previous 
    bullet. """
    app = QtWidgets.QApplication(sys.argv)
    win = MainWindow()
    win.show()
    win.setFocus()
    sys.exit(app.exec_())
